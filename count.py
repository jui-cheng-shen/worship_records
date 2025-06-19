import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
import datetime
import sqlite3

year = str(datetime.datetime.now().year)

filename = year +' record.xlsx'

if not os.path.exists(filename):
    wb = Workbook()
    ws = wb.active
    headers = ["此程式用於記錄聚會的日期、星期、主題、讚美詩、弟兄人數、姊妹人數、慕道者人數和總人數。"]
    ws.append(headers)
    wb.save(filename)

conn = sqlite3.connect(year + ' record.db')
database = conn.cursor()
database.execute('''
    CREATE TABLE IF NOT EXISTS record (
        date TEXT,
        week TEXT,
        topic TEXT,
        hymn1 TEXT,
        hymn2 TEXT,
        brother INTEGER,
        sister INTEGER,
        seeker INTEGER,
        total INTEGER
    )
''')
conn.commit()


record = []

win =tk.Tk()
win.geometry("500x500")
win.title("聚會記錄系統")

group = tk.LabelFrame(win, text="請輸入資料", padx=10, pady=10)
group.pack(padx=10, pady=10)

labels = [
    "日期(MM/DD)", "星期", "主題", "讚美詩1", "讚美詩2", "弟兄人數", "姊妹人數", "慕道者人數", "總人數"
]


record_input = tk.Entry()
entries = []

for label_text in labels:
    frame = tk.Frame(group)
    frame.pack(fill="x", pady=2)
    lbl = tk.Label(frame, text=label_text, width=15, anchor="w")
    lbl.pack(side="left")
    entry = tk.Entry(frame)
    entry.pack(side="left", fill="x", expand=True)
    entries.append(entry)

def add_record():
    data = [entry.get() for entry in entries]

    if any(value.strip() == "" for value in data):
        return
    date_str = data[0].strip()
    current_year = datetime.datetime.now().year

    try:
        if "/" in date_str:
            full_date = f"{current_year}/{date_str}"
            date_obj = datetime.datetime.strptime(full_date, "%Y/%m/%d")
        elif "-" in date_str:
            full_date = f"{current_year}-{date_str}"
            date_obj = datetime.datetime.strptime(full_date, "%Y-%m-%d")
        else:
            messagebox.showerror("date format error", "MM/DD")
            return

        data[0] = date_obj.strftime("%Y/%m/%d")
        entries[0].delete(0, tk.END)
        entries[0].insert(0, data[0])

    except ValueError:
        messagebox.showerror("date format error", "MM/DD")
        return

    try:
        brother = int(data[5])
        sister = int(data[6])
        seeker = int(data[7])
        total = brother + sister + seeker
        data[8] = str(total)
        entries[8].delete(0, tk.END)
        entries[8].insert(0, str(total))
    except ValueError:
        messagebox.showerror("date error", "people must input int")
        return

    record.append(data)
    store_record()
    messagebox.showinfo("success", "success")

    for entry in entries:
        entry.delete(0, tk.END)

def store_record():
    values = [entry.get() for entry in entries]

    date_str = values[0].strip()
    try:
        date_obj = datetime.datetime.strptime(date_str, "%Y/%m/%d")

        month_name = f"{date_obj.year}年{date_obj.month}月"

        wb = load_workbook(filename)

        if month_name not in wb.sheetnames:
            ws = wb.create_sheet(title=month_name)
            headers = [
                "日期", "星期", "主題", "讚美詩1", "讚美詩2", "弟兄人數", "姊妹人數", "慕道者人數", "總人數"
            ]
            ws.append(headers)
        else:
            ws = wb[month_name]

        ws.append(values)
        wb.save(filename)

    except ValueError:
        wb = load_workbook(filename)
        ws = wb.active
        ws.append(values)
        wb.save(filename)

    database.execute('''
        INSERT INTO record (date, week, topic, hymn1, hymn2, brother, sister, seeker, total)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', values)
    conn.commit()

btn = tk.Button(win, text = "儲存", command=add_record)
btn.pack(pady=10)

win.mainloop()
